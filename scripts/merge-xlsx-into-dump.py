#!/usr/bin/env python3
"""
One-off: merge the missing sessions from "Баня Caparica-2.xlsx" into a COPY of
the SQLite dump. Purely additive — the 3 overlapping Jan sessions already in the
dump are left untouched. Operates on a copy; the original dump is never modified.

Usage:
  python3 scripts/merge-xlsx-into-dump.py \
      "/path/to/Баня Caparica-2.xlsx" \
      dumps/banha-dump-20260525-142749.db \
      dumps/banha-merged.db
"""
import sys, re, shutil, sqlite3, calendar, datetime
import openpyxl

XLSX = sys.argv[1] if len(sys.argv) > 1 else "/Users/yevgenmalafeyev/Downloads/Баня Caparica-2.xlsx"
SRC  = sys.argv[2] if len(sys.argv) > 2 else "dumps/banha-dump-20260525-142749.db"
OUT  = sys.argv[3] if len(sys.argv) > 3 else "dumps/banha-merged.db"

DEFAULT_XLSX = 'Время-чай-веник'
DEFAULT_DB   = 'Время, чай, вода'
SUM_HDR = 'Всего, сумма'
STOP = {'Итого', 'Всего внесено', 'Контрольная сумма'}
SKIP_EXP = {'Блюдо 2', 'Блюдо 3', 'Блюдо 4', 'Блюдо 5'}

OVERLAP = {'2026-01-26', '2026-01-19', '2026-01-12'}   # already in dump
SKIP    = {'22.12.2025', '08.12.2025'}                  # empty / zero-cost (skip)
MONTHS  = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']

# explicit participant overrides (decided with the user)
ALIAS_TOKEN = {'сережа': 'сергей'}                      # merge Сережа↔Сергей
NAME_OVERRIDE = {'Люба и ..': 'Люба и Даниил'}          # map to existing dump participant

def canon(s):
    s = s.replace('ё', 'е').replace('Ё', 'Е').strip().lower()
    toks = [ALIAS_TOKEN.get(t, t) for t in re.split(r'[\s,+\-]|(?:\bи\b)', s) if t]
    return ' '.join(sorted(toks))

def parse_date(sheet):
    if re.match(r'\d{4}-\d{2}-\d{2}', sheet):
        return datetime.date.fromisoformat(sheet)
    d, m, y = sheet.split('.')
    return datetime.date(int(y), int(m), int(d))

def epoch(d):  # 18:00 UTC of the session date (day-granular ordering)
    return calendar.timegm((d.year, d.month, d.day, 18, 0, 0, 0, 0, 0))

def parse_sheet(ws):
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    HR = HC = None
    for ri, row in enumerate(grid):
        for ci, v in enumerate(row):
            if isinstance(v, str) and v.strip() == SUM_HDR:
                HR, HC = ri, ci
                break
        if HR is not None:
            break
    if HR is None:
        return None
    cell = lambda r, c: grid[r][c] if r < len(grid) and c < len(grid[r]) else None
    name_row = HR - 1 if (isinstance(cell(HR, HC + 2), str) and cell(HR, HC + 2).startswith('Баньщик')) else HR
    exp_name_col = HC - 1
    parts = []
    c = HC + 2
    while name_row < len(grid) and c < len(grid[name_row]):
        v = cell(name_row, c)
        if isinstance(v, str) and v.strip():
            parts.append((v.strip(), c))
        c += 2
    expenses = []
    for ri in range(HR + 1, len(grid)):
        nm = cell(ri, exp_name_col)
        if nm in (None, ''):
            continue
        nm = str(nm).strip()
        if nm in STOP:
            break
        if nm in SKIP_EXP:
            continue
        total = cell(ri, HC)
        total = float(total) if isinstance(total, (int, float)) else 0.0
        shares = {}
        for (p, ci) in parts:
            v = cell(ri, ci)
            if isinstance(v, (int, float)) and v not in (0, 0.0):
                shares[p] = float(v)
        # skip empty leftover rows: no cost and nobody assigned
        if total == 0 and not shares:
            continue
        expenses.append({'name': nm, 'total': total, 'shares': shares})
    defexp = next((e for e in expenses if e['name'] == DEFAULT_XLSX), None)
    pcount = {p: int(round(defexp['shares'].get(p, 1))) if defexp else 1 for p, _ in parts}
    for p in pcount:
        if pcount[p] < 1:
            pcount[p] = 1
    return {'participants': [p for p, _ in parts], 'person_count': pcount, 'expenses': expenses}

# ---- load xlsx ----
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
sheets = [n for n in wb.sheetnames if n != 'Template']
to_import = sorted([n for n in sheets if n not in OVERLAP and n not in SKIP], key=parse_date)
parsed = {n: parse_sheet(wb[n]) for n in to_import}

# ---- copy dump and connect ----
shutil.copyfile(SRC, OUT)
con = sqlite3.connect(OUT)
con.row_factory = sqlite3.Row
con.execute("PRAGMA foreign_keys = OFF")

def counts():
    return {t: con.execute(f"SELECT count(*) c FROM {t}").fetchone()['c'] for t in
            ['sessions', 'participants', 'session_participants', 'expenses',
             'expense_assignments', 'session_participant_meta', 'expense_templates']}

before = counts()

dump_names = [r['name'] for r in con.execute("SELECT id,name FROM participants")]
dump_by_canon = {}
for n in dump_names:
    dump_by_canon.setdefault(canon(n), n)

created_new = []
existing_id_by_name = {}   # dump participant name -> id
new_id_by_canon = {}       # canonical key -> id (for newly created participants)

# When several xlsx spellings map to the same new participant, prefer an
# 'и'-form display name (e.g. "Люда и Сергей" over "Сережа и Люда").
new_disp_by_canon = {}
for n in to_import:
    for p in parsed[n]['participants']:
        tgt = NAME_OVERRIDE.get(p, p)
        if tgt in dump_names:
            continue
        ck = canon(tgt)
        if ck in dump_by_canon:
            continue
        cur = new_disp_by_canon.get(ck)
        if cur is None or (' и ' in tgt and ' и ' not in cur):
            new_disp_by_canon[ck] = tgt

def existing_id(name):
    if name not in existing_id_by_name:
        existing_id_by_name[name] = con.execute(
            "SELECT id FROM participants WHERE name=?", (name,)).fetchone()['id']
    return existing_id_by_name[name]

def participant_id(xlsx_name):
    """Resolve an xlsx participant name to a participants.id: reuse an existing
    dump participant (exact / canonical / explicit override) or create a new one
    (deduped by canonical key across the import)."""
    target = NAME_OVERRIDE.get(xlsx_name, xlsx_name)
    if target in dump_names:
        return existing_id(target)
    ck = canon(target)
    if ck in dump_by_canon:
        return existing_id(dump_by_canon[ck])
    if ck in new_id_by_canon:
        return new_id_by_canon[ck]
    disp = new_disp_by_canon.get(ck, target)
    row = con.execute("SELECT id FROM participants WHERE name=?", (disp,)).fetchone()
    if row:
        pid = row['id']
    else:
        pid = con.execute(
            "INSERT INTO participants (name, activity_score, created_at) VALUES (?,?,?)",
            (disp, 0, calendar.timegm(datetime.datetime.utcnow().timetuple()))
        ).lastrowid
        created_new.append(disp)
    new_id_by_canon[ck] = pid
    return pid

# ---- insert sessions ----
for n in to_import:
    s = parsed[n]
    d = parse_date(n)
    ts = epoch(d)
    name = f"{d.day} {MONTHS[d.month - 1]} {d.year}"
    sid = con.execute(
        "INSERT INTO sessions (name, hidden, duty_person, created_at, updated_at) VALUES (?,?,?,?,?)",
        (name, 0, None, ts, ts)
    ).lastrowid

    spid = {}   # xlsx participant name -> session_participant id
    for p in s['participants']:
        pid = participant_id(p)
        sp = con.execute(
            "INSERT INTO session_participants (session_id, participant_id, person_count) VALUES (?,?,?)",
            (sid, pid, s['person_count'].get(p, 1))
        ).lastrowid
        spid[p] = sp
        con.execute(
            "INSERT INTO session_participant_meta (session_participant_id, has_paid, joined_at) VALUES (?,?,?)",
            (sp, 0, ts)
        )

    for e in s['expenses']:
        ename = DEFAULT_DB if e['name'] == DEFAULT_XLSX else e['name']
        item_count = 1 if e['name'] == DEFAULT_XLSX else max(1, int(round(sum(e['shares'].values()))) or 1)
        eid = con.execute(
            "INSERT INTO expenses (session_id, name, item_count, total_cost) VALUES (?,?,?,?)",
            (sid, ename, item_count, e['total'])
        ).lastrowid
        for pname, share in e['shares'].items():
            if pname in spid:
                con.execute(
                    "INSERT INTO expense_assignments (expense_id, session_participant_id, share) VALUES (?,?,?)",
                    (eid, spid[pname], share)
                )

con.commit()
after = counts()

print(f"Imported {len(to_import)} sessions into {OUT}\n")
print(f"{'table':<28}{'before':>8}{'after':>8}{'added':>8}")
for t in before:
    print(f"{t:<28}{before[t]:>8}{after[t]:>8}{after[t]-before[t]:>8}")
print(f"\nNew participants created ({len(created_new)}):")
for nm in sorted(set(created_new)):
    print(f"  {nm}")
con.close()
